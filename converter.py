from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import tkinter
import tkinter.messagebox
from tkinter import *

#district list
DList = ['BANKURA','BIRBHUM','BURDWAN','COOCHBEHAR','DAKSHIN DINAJPUR','DARJEELING','HOOGHLY','HOWRAH','JALPAIGURI','MALDA','MURSHIDABAD','NADIA','NORTH 24 PARGANAS','PASCHIM MEDINIPUR','PURBA MEDINIPUR','PURULIA','SOUTH 24 PARGANAS','UTTAR DINAJPUR']
CropList = ['AUS','AMAN','BORO','WHEAT','MAIZE','JUTE','MUSUR','MASKALAI','KHESARI','GRAM','MUSTARD','TIL','POTATO','SUGARCANE']
years = []

def getList():
    lis = ent.get()
    global DList, NList, years
    year = enty.get()
    yr = enty.get()
    if yr != '':
        years = []
        for wrd in (yr).split(','):
            years.append(wrd)
    else:
        tkinter.messagebox.showerror(title="Error",message="No year entered.\nPlease enter years.")
    if lis != '':
        DList = []
        for wrd in (lis).split(','):
            DList.append(wrd)
    root.destroy()

root = tkinter.Tk()
root.wm_title("Convertor")
lab = Label(root, width=15, text="District names", anchor='w')
ent = Entry(root, width=30)
laby = Label(root, width=15, text="Year", anchor='w')
enty = Entry(root, width=30)
inst = Label(root, justify=LEFT, fg='#656565', text="Enter district names and years separated by commas.\nEnter nothing and press okay to convert\nall districts.\nSeparate year with \'-\'")
but = Button(root, width=8, text="Okay", command=getList)
lab.grid(row=1,column=1)
laby.grid(row=2,column=1)
ent.grid(row=1,column=2,columnspan=2)
enty.grid(row=2,column=2,columnspan=2)
inst.grid(row=3,column=1,columnspan=2)
but.grid(row=3,column=3)
root.mainloop()

for year in years:

    if not os.path.exists("OUTPUT\\"+year):
        os.makedirs("OUTPUT\\"+year)
    
    for dist in DList:
        
        wb = Workbook()
        ws = wb.active
        ws.title = "18.1"

        ws.cell(row=1, column=1).value="Area, Production and Yield rates of Major Crops in the Blocks of "+dist+" for the year "+year
        ws.cell(row=1, column=1).font = Font(bold=True,size=13)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=17)
        
        c=0
        r=5
        
        for i in range (1,32):
            ws.cell(row=r+i,column=1).value=i
            
        for i in range(0,14):
            crop = CropList[i]
            if (i%5 == 0):
                if i!=0:
                    c+=3
                ws.merge_cells(start_row=2,start_column=c+1,end_row=2,end_column=c+17)
                ws.merge_cells(start_row=3,start_column=c+1,end_row=4,end_column=c+1)
                ws.merge_cells(start_row=3,start_column=c+2,end_row=4,end_column=c+2)
                ws.cell(row=2, column=c+1).value="TABLE 18.1"
                ws.cell(row=3, column=c+1).value="Sl. no"
                ws.cell(row=3, column=c+2).value="Name of Block"
                for j in range (1,18):
                    ws.cell(row=5,column=c+j).value = "("+str(j)+")"
                    
            c+=3

            ws.merge_cells(start_row=3,start_column=c,end_row=3,end_column=c+2)
            ws.cell(row=3, column=c).value=crop
            ws.cell(row=4,column=c).value="Area"
            ws.cell(row=4,column=c+1).value="Prod."
            ws.cell(row=4,column=c+2).value="Yield"


            filepath = Path("INPUT\\"+year+"\\"+dist.replace(' ','_')+"\\"+dist.replace(' ','_')+"_"+crop+".txt")
            if filepath.is_file():
                f = open ("INPUT\\"+year+"\\"+dist.replace(' ','_')+"\\"+dist.replace(' ','_')+"_"+crop+".txt",'r')
                for line in f:
                    if (line != "\n"):
                        if ((line[2]>='0') & (line[2]<='9')):
                            wno=0
                            for word in line.split():
                                wno+=1
                                if (wno == 4):
                                    j = int(word)
                                    ws.cell(row=r+j,column=19).value = j
                                    ws.cell(row=r+j,column=37).value = j
                                if (wno == 5):
                                    if ((word[0]>='0') & (word[0]<='9')):
                                        ws.cell(row=r+j,column=c+2).value = "ERROR"
                                        ws.cell(row=r+j,column=c+1).value = "ERROR"
                                        ws.cell(row=r+j,column=c).value = "ERROR"
                                        break
                                    else:
                                        nm=word
                                if (wno == 6):
                                    if ((word[0]<'0') | (word[0]>'9')):
                                        nm = nm+" "+word
                                        wno=wno-1
                                    else:
                                        ws.cell(row=r+j,column=2).value = nm
                                        ws.cell(row=r+j,column=20).value = nm
                                        ws.cell(row=r+j,column=38).value = nm
                                if (crop == 'AUS')|(crop == 'AMAN')|(crop == 'BORO'):
                                    if (wno ==  10):
                                        ws.cell(row=r+j,column=c+2).value = float(word)
                                    if (wno == 8):
                                        ws.cell(row=r+j,column=c).value = float(word)
                                    if (wno == 11):
                                        ws.cell(row=r+j,column=c+1).value = float(word)
                                else:
                                    if (wno == 7):
                                        ws.cell(row=r+j,column=c+2).value = float(word)
                                    if (wno == 8):
                                        ws.cell(row=r+j,column=c).value = float(word)
                                    if (wno == 9):
                                        ws.cell(row=r+j,column=c+1).value = float(word)
                f.close()

        wb.save("OUTPUT\\"+year+"\\"+dist.replace(" ","_")+"_Crop_"+year+".xlsx")
